from flask import Flask,request,jsonify
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
import time


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:root@localhost:5432/Neet'
db = SQLAlchemy(app)

class HashTable(db.Model):
    __tablename__ ="HashTable"
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    mark = db.Column(db.String(200))
    range = db.Column(db.String(200))
    year = db.Column(db.String(200))
    is_active = db.Column(db.Boolean)

@app.route('/')
def start():
    return {'Afsar': 'Khan'}

@app.route('/Demo', methods=['POST'])
def game():
    if request.method=="POST":
        file=request.files['file']
        workbook = load_workbook(file)
        a=[]
        sheet1=[]
        for sheet_name in workbook.sheetnames:
            a.append(sheet_name)

        for sheet_name in workbook.sheetnames:
            if a[0]==sheet_name:
                a[0]=workbook[sheet_name]
                for row in a[0].iter_rows(min_row=2, values_only=True):
                    sheet1.append(row)
        print(sheet1)
        for mark_, range_, year_, is_active_ in sheet1:
            data2 = HashTable(mark=mark_, range=range_, year=year_, is_active=is_active_)
            db.session.add(data2)
        db.session.commit()
    return jsonify({'msg':'data inserted'})

if __name__ == '__main__':
    app.run(debug=True)

